import os
import pandas as pd
from openpyxl import load_workbook

# Function to process a single .Xlsx file and return a modified DataFrame
def Xlsx_to_csv(Xlsx_path):
    # Load the Excel file
    workbook = load_workbook(Xlsx_path, data_only=True)
    sheet = workbook.active

    # Read the main table (A15 to the last filled row)
    headers = [sheet.cell(row=15, column=col).value for col in range(1, 8)]
    data = []
    row = 16
    while True:
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, 8)]
        if all(cell is None for cell in row_data):  # Check if the entire row is empty
            break
        data.append(row_data)
        row += 1

    # Create the DataFrame
    df = pd.DataFrame(data, columns=headers)

    # Remove the last row (sum row which is not needed)
    df = df.iloc[:-1]

    # Read the value from cell D9
    valor_d9 = sheet['D9'].value  # Example: "AP7-129+500-100280000000"

    # Create the 3 new columns based on the value in D9
    etd = valor_d9
    via = valor_d9.split('-')[0]
    pk = float(valor_d9.split('-')[1].replace('+', '.'))

    # Add the new columns at the beginning of the DataFrame
    df.insert(0, 'ETD', etd)
    df.insert(1, 'via', via)
    df.insert(2, 'pk', pk)

    # Convert columns from column E onwards to integers
    for col in df.columns[4:]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    # Read the lane code mapping file (containing ETD, CARRIL, and SENTIDO columns)
    codis_carrils_df = pd.read_excel("../files/carril_code.xlsx", usecols=['ETD', 'CARRIL', 'SENTIDO'])

    # Add the SENTIDO column by merging with ETD and CARRIL
    df = df.merge(codis_carrils_df, on=['ETD', 'CARRIL'], how='left')

    # Insert the 'SENTIDO' column at the desired position (after column E)
    col_posicio = 5  # Column F (index 5 in Python)
    sentit_col = df.pop('SENTIDO')  # Remove the 'SENTIDO' column
    df.insert(col_posicio, 'SENTIDO', sentit_col)  # Insert it at the desired position

    # Define the vector of column names
    noms_columnes = ['etd', 'via', 'pk', 'dat', 'carr', 'sen', 
                     'vehicles_lleugers', 'camions_rigids', 
                     'camions_articulats', 'autobusos', 'total']

    # Replace the column names
    df.columns = noms_columnes
    
    return df

# Main function to traverse directories and files
def manage_directories(origin, destination):
    for root, dirs, files in os.walk(origin):
        # Compute the new path to replicate the folder structure
        rel_path = os.path.relpath(root, origin)
        dest_path = os.path.join(destination, rel_path)
        os.makedirs(dest_path, exist_ok=True)

        # Process each .Xlsx file
        for file in files:
            if file.endswith('.Xlsx'):
                Xlsx_path = os.path.join(root, file)
                print(f"Processing: {Xlsx_path}")
                
                # Process the Excel file
                df_modificat = Xlsx_to_csv(Xlsx_path)
                
                # Save the result as a .csv file
                nom_csv = os.path.splitext(file)[0] + '.csv'
                output_path = os.path.join(dest_path, nom_csv)
                df_modificat.to_csv(output_path, index=False, sep=';', encoding='utf-8-sig')
                print(f"Saved: {output_path}")
